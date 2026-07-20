from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[3]
OUTPUT_DIR = BASE_DIR / "generated_quotes"

HEADER_FILL = PatternFill(
    start_color="1E293B", end_color="1E293B", fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
LABEL_FONT = Font(size=9, color="5B6472")
BOLD_FONT = Font(bold=True)


def _autosize_columns(worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def generate_quote_excel(quote: dict) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    quote_id = quote["quote_id"]
    output_path = OUTPUT_DIR / f"{quote_id}.xlsx"

    workbook = Workbook()

    # --- Summary sheet ---
    summary = workbook.active
    summary.title = "Summary"

    summary["A1"] = "QuoteIQ"
    summary["A1"].font = TITLE_FONT
    summary["C1"] = f"Quote {quote_id}"
    summary["C1"].font = BOLD_FONT

    rows = [
        ("Customer", quote.get("customer_name")),
        ("Customer ID", quote.get("customer_id")),
        ("Price class", quote.get("price_class")),
        ("Quote status", quote.get("quote_status")),
        ("Quote confidence (%)", quote.get("quote_confidence")),
        ("Subtotal ($)", quote.get("quote_subtotal")),
        ("Estimated margin (%)", quote.get("estimated_margin_pct")),
        ("Risk count", quote.get("risk_count")),
        ("Review required", "Yes" if quote.get("review_required") else "No"),
        ("Created by", quote.get("created_by")),
        ("Approved by", quote.get("approved_by")),
        ("Sales order ID", quote.get("sales_order_id")),
    ]

    for row_index, (label, value) in enumerate(rows, start=3):
        summary.cell(row=row_index, column=1, value=label).font = LABEL_FONT
        summary.cell(row=row_index, column=2, value=value)

    _autosize_columns(summary, [22, 28, 4, 20])

    # --- Line items sheet ---
    lines_sheet = workbook.create_sheet("Line Items")

    headers = [
        "SKU",
        "Product name",
        "Category",
        "Description (as requested)",
        "Quantity",
        "UOM",
        "Base cost",
        "List price",
        "Unit price",
        "Line total",
        "Margin %",
        "Pricing rule",
        "Risk flag",
        "Risk reason",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = lines_sheet.cell(row=1, column=col_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_index, line in enumerate(quote.get("priced_lines", []), start=2):
        lines_sheet.cell(row=row_index, column=1, value=line.get("sku"))
        lines_sheet.cell(row=row_index, column=2, value=line.get("product_name"))
        lines_sheet.cell(row=row_index, column=3, value=line.get("category"))
        lines_sheet.cell(row=row_index, column=4, value=line.get("description_raw"))
        lines_sheet.cell(row=row_index, column=5, value=line.get("quantity"))
        lines_sheet.cell(row=row_index, column=6, value=line.get("uom_raw"))
        lines_sheet.cell(row=row_index, column=7, value=line.get("base_cost"))
        lines_sheet.cell(row=row_index, column=8, value=line.get("list_price"))
        lines_sheet.cell(row=row_index, column=9, value=line.get("unit_price"))
        lines_sheet.cell(row=row_index, column=10, value=line.get("line_total"))
        lines_sheet.cell(row=row_index, column=11, value=line.get("margin_pct"))
        lines_sheet.cell(row=row_index, column=12, value=line.get("pricing_rule_applied"))
        lines_sheet.cell(
            row=row_index, column=13, value="Yes" if line.get("risk_flag") else "No"
        )
        lines_sheet.cell(row=row_index, column=14, value=line.get("risk_reason"))

    _autosize_columns(
        lines_sheet,
        [16, 26, 14, 32, 9, 8, 11, 11, 11, 12, 10, 22, 10, 32],
    )
    lines_sheet.freeze_panes = "A2"

    workbook.save(output_path)

    return output_path